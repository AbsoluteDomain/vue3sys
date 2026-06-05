"""系统管理-用户视图。

"""

from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, extend_schema, inline_serializer
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from io import BytesIO
import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from django.utils.dateparse import parse_datetime

from core.exceptions.business import BusinessException
from core.permissions.decorators import permission_required
from core.permissions.perms import HasPerm
from apps.system.dept.models import Department
from apps.system.roles.models import Role
from apps.system.users.models import User
from core.permissions.data_scope import data_permission_required, DataPermission
from core.response import error, success
from apps.auth.utils.session_utils import invalidate_user_sessions
from apps.system.utils.email_utils import generate_and_send_email_code
from apps.system.utils.mobile_utils import send_mobile_code
from apps.system.utils.rate_limit import ip_rate_limit
from core.openapi import page_resp, resp
from core.viewsets import BaseModelViewSet
from .filters import UserFilter
from .serializers import (
    UserInfoSerializer, UserPageSerializer, UserCreateSerializer, 
    UserFormSerializer, UserUpdateSerializer, UserPasswordResetSerializer,
    UserStatusSerializer, UserProfileUpdateSerializer, UserPasswordChangeSerializer,
    UserEmailUpdateSerializer, DepartmentOptionSerializer, UserMobileUpdateSerializer, PasswordVerifySerializer
)
from apps.system.utils.decorators import log_api
import re
import logging

_logger = logging.getLogger(__name__)

def _mask_mobile(mobile: str | None) -> str | None:
    """手机号脱敏（仅用于日志输出）。"""
    if not mobile:
        return mobile
    m = str(mobile)
    if len(m) < 7:
        return "***"
    return f"{m[:3]}****{m[-4:]}"

def _mask_email(email: str | None) -> str | None:
    """邮箱脱敏（仅用于日志输出）。"""
    if not email:
        return email
    e = str(email)
    if "@" not in e:
        return "***"
    name, domain = e.split("@", 1)
    name_masked = (name[:2] + "***") if len(name) >= 2 else "***"
    return f"{name_masked}@{domain}"

def build_user_page_query(*, keywords: str, status, dept_id, role_ids: str, create_time: str, viewable_users):
    """构建用户分页查询条件。

    - **用途**：将前端分页查询参数转换为 Django ORM 的 `Q` 条件。
    - **约束**：默认只查询 `is_deleted=0` 用户；若 `viewable_users != "all"`，按可见用户集合做二次过滤。
    - **create_time**：期望为 `start,end` 两段（可被 `parse_datetime` 解析）。
    """
    query = Q(is_deleted=0)

    if viewable_users != "all":
        query &= Q(id__in=viewable_users)

    if keywords:
        query &= Q(username__icontains=keywords) | Q(nickname__icontains=keywords) | Q(mobile__icontains=keywords)

    if status is not None and status != "":
        query &= Q(status=int(status))

    if dept_id:
        query &= Q(dept_id=dept_id)

    if role_ids:
        role_id_list = role_ids.split(",")
        query &= Q(roles__id__in=role_id_list)

    # root 用户（ROOT 角色）不在用户列表/导出中展示
    query &= ~Q(roles__code="ROOT")

    if create_time:
        time_range = create_time.split(",")
        if len(time_range) == 2:
            start_time = parse_datetime(time_range[0])
            end_time = parse_datetime(time_range[1])
            if start_time and end_time:
                query &= Q(date_joined__gte=start_time, date_joined__lte=end_time)

    return query


def _check_user_data_permission(*, request, operation: str, target_user_id: int) -> bool:
    """用户对象级数据权限校验。"""
    user = getattr(request, "user", None)
    return bool(data_permission_required(user, operation, target_user_id))

def delete_users(*, ids: str, current_user_id: int, current_user):
    """批量逻辑删除用户。

    - **ids**：英文逗号分隔的用户ID字符串。
    - **current_user_id**：当前登录用户ID；禁止删除自身。
    - **current_user**：当前登录用户对象，用于数据权限校验。
    - **返回**：`{"deleted_count": <int>}` 或 `{"error": <str>, "status": <int>}`。
    """
    _logger.info("delete_users start current_user_id=%s ids=%s", current_user_id, ids)
    try:
        id_list = [int(i.strip()) for i in str(ids).split(",") if i.strip()]
    except ValueError:
        _logger.warning("delete_users invalid ids format current_user_id=%s ids=%s", current_user_id, ids)
        raise BusinessException("用户 ID 格式不正确", code="A0400", status=400)

    if not id_list:
        _logger.warning("delete_users empty ids current_user_id=%s", current_user_id)
        raise BusinessException("未指定要删除的用户 ID", code="A0400", status=400)

    if current_user_id in id_list:
        _logger.warning("delete_users attempted to delete self current_user_id=%s", current_user_id)
        raise BusinessException("不能删除当前登录的用户", code="A0400", status=400)

    authorized_ids = []
    unauthorized_ids = []
    for user_id in id_list:
        has_permission = data_permission_required(current_user, "delete", user_id)
        if has_permission:
            authorized_ids.append(user_id)
        else:
            unauthorized_ids.append(user_id)

    if unauthorized_ids:
        _logger.warning(
            "delete_users forbidden current_user_id=%s unauthorized_ids=%s", current_user_id, unauthorized_ids
        )
        raise BusinessException(
            f"您没有权限删除以下用户: {','.join(map(str, unauthorized_ids))}",
            code="A0301",
            status=403,
        )

    users = User.objects.filter(id__in=authorized_ids, is_deleted=0)
    found_ids = list(users.values_list("id", flat=True))
    not_found_ids = set(authorized_ids) - set(found_ids)
    if not_found_ids:
        _logger.warning(
            "delete_users not_found current_user_id=%s not_found_ids=%s", current_user_id, list(not_found_ids)
        )
        raise BusinessException(
            f"用户 ID {','.join(map(str, not_found_ids))} 不存在或已删除",
            code="A0400",
            status=400,
        )

    users.update(is_deleted=1)
    # 强制被删除的用户下线
    for user_id in authorized_ids:
        invalidate_user_sessions(user_id)
    _logger.info("delete_users success current_user_id=%s deleted_count=%s", current_user_id, len(found_ids))
    return {"deleted_count": len(found_ids)}

def build_user_import_template_excel_bytes():
    """构建用户导入 Excel 模板（返回 bytes）。"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "用户导入"

    headers = ["用户名", "昵称", "性别", "手机号", "邮箱", "角色", "部门"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()

def import_users_from_excel(*, file):
    """从 Excel 导入用户。

    - **file**：上传文件对象（可直接被 openpyxl 读取）。
    - **表头**：要求包含 `用户名/昵称/性别/手机号/邮箱/角色/部门`。
    - **返回结构**：保持与前端约定一致：
      - `data.validCount/invalidCount/messageList`
      - `msg` 汇总信息
    """
    _logger.info("import_users_from_excel start")
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
    except Exception as e:
        _logger.exception("import_users_from_excel openpyxl load_workbook failed")
        raise BusinessException(f"读取 Excel 文件失败: {str(e)}", code="A0400", status=400)

    required_columns = ["用户名", "昵称", "性别", "手机号", "邮箱", "角色", "部门"]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    header_row = list(header_row or [])
    header_index = {str(name).strip(): idx for idx, name in enumerate(header_row) if name is not None}

    missing_columns = [col for col in required_columns if col not in header_index]
    if missing_columns:
        _logger.warning("import_users_from_excel missing_columns=%s", missing_columns)
        raise BusinessException(f"缺少必填列: {', '.join(missing_columns)}", code="A0400", status=400)

    def _get_cell_value(row_values, key: str):
        idx = header_index.get(key)
        if idx is None or idx >= len(row_values):
            return None
        return row_values[idx]

    def _to_opt_str(v):
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    # 构建部门映射（同时支持名称和编码）
    dept_map = {}
    for d in Department.objects.all():
        dept_map[d.name] = d
        if d.code:
            dept_map[d.code] = d
    # 构建角色映射（同时支持名称和编码）
    role_map = {}
    for r in Role.objects.all():
        role_map[r.name] = r
        if r.code:
            role_map[r.code] = r
    existing_usernames = set(User.objects.values_list("username", flat=True))

    valid_count = 0
    invalid_count = 0
    message_list = []

    _logger.info(
        "import_users_from_excel parsed_header columns=%s data_rows=%s",
        list(header_index.keys()),
        max((worksheet.max_row or 1) - 1, 0),
    )

    for index, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
        try:
            username = _to_opt_str(_get_cell_value(row_values, "用户名")) or ""
            nickname = _to_opt_str(_get_cell_value(row_values, "昵称")) or ""

            gender_value = _get_cell_value(row_values, "性别")
            if gender_value is None or str(gender_value).strip() == "":
                gender = 1
            else:
                gender_str = str(gender_value).strip()
                if gender_str in ["0", "保密"]:
                    gender = 0
                elif gender_str in ["1", "男"]:
                    gender = 1
                elif gender_str in ["2", "女"]:
                    gender = 2
                else:
                    try:
                        gender = int(gender_str)
                    except ValueError:
                        message_list.append(
                            f"第{index + 2}行 性别值\"{gender_str}\"不正确，应为 0(保密)/1(男)/2(女)"
                        )
                        invalid_count += 1
                        continue

            mobile = _to_opt_str(_get_cell_value(row_values, "手机号"))
            email = _to_opt_str(_get_cell_value(row_values, "邮箱"))
            role_names = _to_opt_str(_get_cell_value(row_values, "角色"))
            dept_name = _to_opt_str(_get_cell_value(row_values, "部门"))

            if not username or not nickname:
                _logger.warning("import_users_from_excel invalid_row row=%s reason=empty_username_or_nickname", index + 2)
                message_list.append(f"第{index + 2}行 用户名和昵称不能为空")
                invalid_count += 1
                continue

            if username in existing_usernames:
                _logger.warning("import_users_from_excel invalid_row row=%s reason=username_exists", index + 2)
                message_list.append(f"第{index + 2}行 用户名\"{username}\"已存在")
                invalid_count += 1
                continue

            if mobile and not re.match(r"^1[3-9]\d{9}$", mobile):
                _logger.warning(
                    "import_users_from_excel invalid_row row=%s reason=invalid_mobile mobile=%s",
                    index + 2,
                    _mask_mobile(mobile),
                )
                message_list.append(f"第{index + 2}行 手机号\"{mobile}\"格式不正确")
                invalid_count += 1
                continue

            if email and not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email):
                _logger.warning(
                    "import_users_from_excel invalid_row row=%s reason=invalid_email email=%s",
                    index + 2,
                    _mask_email(email),
                )
                message_list.append(f"第{index + 2}行 邮箱\"{email}\"格式不正确")
                invalid_count += 1
                continue

            if gender not in [0, 1, 2]:
                message_list.append(f"第{index + 2}行 性别值\"{gender}\"不正确，应为 0(保密)/1(男)/2(女)")
                invalid_count += 1
                continue

            dept = None
            if dept_name:
                dept = dept_map.get(dept_name)
                if dept is None:
                    message_list.append(f"第{index + 2}行 部门\"{dept_name}\"不存在")
                    invalid_count += 1
                    continue

            roles = []
            if role_names:
                role_name_list = [name.strip() for name in role_names.split(",") if name.strip()]
                missing_roles = [name for name in role_name_list if name not in role_map]
                if missing_roles:
                    for role_name in missing_roles:
                        message_list.append(f"第{index + 2}行 角色\"{role_name}\"不存在")
                    invalid_count += 1
                    continue
                roles = [role_map[name] for name in role_name_list]

            user = User.objects.create(
                username=username,
                nickname=nickname,
                gender=gender,
                mobile=mobile,
                email=email,
                dept=dept,
                status=1,
                is_deleted=0,
            )
            user.set_password("123456")
            user.save()

            if roles:
                user.roles.set(roles)

            existing_usernames.add(username)
            valid_count += 1
        except Exception as e:
            _logger.exception("import_users_from_excel unexpected_row_error row=%s", index + 2)
            message_list.append(f"第{index + 2}行 导入失败 - {str(e)}")
            invalid_count += 1

    return {
        "data": {
            "validCount": valid_count,
            "invalidCount": invalid_count,
            "messageList": message_list,
        },
        "msg": f"导入完成，成功导入{valid_count}条数据，失败{invalid_count}条",
    }

def export_users_to_excel_bytes(*, keywords: str, status, dept_id, role_ids: str, create_time: str, field: str, direction: str):
    """导出用户列表为 Excel（返回 bytes）。

    - **过滤参数**：与分页查询一致（keywords/status/deptId/roleIds/createTime）。
    - **排序**：`field + direction(ASC/DESC)`。
    - **返回**：Excel 文件内容（bytes），由上层 HttpResponse 作为附件返回。
    """
    query = build_user_page_query(
        keywords=keywords,
        status=status,
        dept_id=dept_id,
        role_ids=role_ids,
        create_time=create_time,
        viewable_users="all",
    )

    order_field = field if (direction or "DESC").upper() == "ASC" else f"-{field}"
    users = User.objects.filter(query).distinct().order_by(order_field)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "用户列表"

    headers = ["用户名", "昵称", "部门", "性别", "手机号", "邮箱", "创建时间"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    for row_num, user in enumerate(users, 2):
        dept_name = user.dept.name if user.dept else ""
        gender_text = ""
        if user.gender == 0:
            gender_text = "保密"
        elif user.gender == 1:
            gender_text = "男"
        elif user.gender == 2:
            gender_text = "女"
        create_time_str = user.date_joined.strftime("%Y/%m/%d %H:%M:%S") if user.date_joined else ""

        worksheet.cell(row=row_num, column=1, value=user.username)
        worksheet.cell(row=row_num, column=2, value=user.nickname)
        worksheet.cell(row=row_num, column=3, value=dept_name)
        worksheet.cell(row=row_num, column=4, value=gender_text)
        worksheet.cell(row=row_num, column=5, value=user.mobile)
        worksheet.cell(row=row_num, column=6, value=user.email)
        worksheet.cell(row=row_num, column=7, value=create_time_str)

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column].width = adjusted_width

    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()


@extend_schema(tags=["02.用户接口"])
class UserViewSet(BaseModelViewSet):
    """系统管理-用户ViewSet。"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasPerm]
    queryset = User.objects.filter(is_deleted=0)
    filterset_class = UserFilter

    def get_queryset(self):
        """按 action 动态构造 queryset。

        - `page`：支持关键词/状态/部门/角色/创建时间范围过滤，并应用数据权限。
        - 其他 action：默认仅查询未删除用户。
        """
        qs = User.objects.filter(is_deleted=0)

        if getattr(self, "action", None) in ["list"]:
            field = self.request.query_params.get("field", "id")
            direction = self.request.query_params.get("direction", "DESC")

            viewable_users = data_permission_required(getattr(self.request, "user", None), "select")

            query = Q(is_deleted=0)
            if viewable_users != "all":
                query &= Q(id__in=viewable_users)

            order_field = field if (direction or "DESC").upper() == "ASC" else f"-{field}"
            return qs.filter(query).distinct().order_by(order_field)

        return qs

    def get_serializer_class(self):
        """根据 action 返回对应序列化器。"""
        if self.action in ["list"]:
            return UserPageSerializer
        if self.action == "create":
            return UserCreateSerializer
        if self.action in ["update", "partial_update"]:
            return UserUpdateSerializer
        if self.action == "form":
            return UserFormSerializer
        return UserFormSerializer

    @permission_required(["sys:user:list"])
    @extend_schema(
        summary="用户分页列表",
        parameters=[
            OpenApiParameter(
                name="keywords",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="关键词（用户名/昵称/手机号）",
            ),
            OpenApiParameter(
                name="status",
                location=OpenApiParameter.QUERY,
                required=False,
                type=int,
                description="用户状态：1 启用 / 0 禁用",
            ),
            OpenApiParameter(
                name="deptId",
                location=OpenApiParameter.QUERY,
                required=False,
                type=int,
                description="部门ID",
            ),
            OpenApiParameter(
                name="roleIds",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="角色ID列表，多个以英文逗号(,)分割",
            ),
            OpenApiParameter(
                name="createTime",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="创建时间范围：start,end（可被 parse_datetime 解析）",
            ),
            OpenApiParameter(
                name="field",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="排序字段（默认 id）",
            ),
            OpenApiParameter(
                name="direction",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="排序方向：ASC / DESC（默认 DESC）",
            ),
            OpenApiParameter(
                name="pageNum",
                location=OpenApiParameter.QUERY,
                required=False,
                type=int,
                description="页码（默认 1）",
            ),
            OpenApiParameter(
                name="pageSize",
                location=OpenApiParameter.QUERY,
                required=False,
                type=int,
                description="每页记录数（默认 10）",
            ),
        ],
        responses=page_resp("UserPageResult", UserPageSerializer(many=True)),
    )
    @DataPermission(dept_field="dept_id", user_field="create_by")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @permission_required(["sys:user:create"])
    @extend_schema(
        summary="新增用户",
        request=UserCreateSerializer,
        responses=resp("UserCreateResult", UserCreateSerializer()),
    )
    def create(self, request, *args, **kwargs):
        """新增用户；成功后会失效当前用户会话缓存（避免权限/角色变更延迟生效）。"""
        serializer = UserCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            invalidate_user_sessions(request.user.id)
            return success(serializer.data, msg="用户创建成功")
        return error(serializer.errors)

    @permission_required(["sys:user:update"])
    @extend_schema(
        summary="用户详情",
        responses=resp("UserDetailResult", UserFormSerializer()),
    )
    def retrieve(self, request, *args, **kwargs):
        """用户详情（用于编辑回显）。"""
        user = self.get_object()
        serializer = UserFormSerializer(user)
        return success(serializer.data, msg="获取用户信息成功")

    @permission_required(["sys:user:update"])
    @extend_schema(
        summary="更新用户信息",
        request=UserUpdateSerializer,
        responses=resp("UserUpdateResult", UserUpdateSerializer()),
    )
    def update(self, request, *args, **kwargs):
        user = self.get_object()
        if not _check_user_data_permission(request=request, operation="update", target_user_id=user.id):
            raise PermissionDenied("您没有权限更新该用户信息")

        serializer = UserUpdateSerializer(user, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return success(serializer.data, msg="更新用户信息成功")
        return error(serializer.errors)

    @permission_required(["sys:user:delete"])
    @extend_schema(
        summary="删除用户",
        parameters=[
            OpenApiParameter(
                name="ids",
                location=OpenApiParameter.PATH,
                required=True,
                type=str,
                description="用户ID，多个以英文逗号(,)分割",
            )
        ],
        responses=resp("UserDeleteResult", serializers.DictField(required=False)),
    )
    def destroy(self, request, *args, **kwargs):
        """批量删除用户（逻辑删除）。

        - 路径参数 `ids` 支持英文逗号分隔。
        - 内部会调用 `delete_users` 并执行数据权限校验。
        """
        ids = kwargs.get(self.lookup_url_kwarg or self.lookup_field)
        result = delete_users(ids=str(ids), current_user_id=request.user.id, current_user=request.user)
        return success({}, msg=f"成功删除 {result['deleted_count']} 个用户")

    @log_api('获取当前用户信息', module='用户模块')
    @extend_schema(
        summary="获取当前用户信息",
        responses=resp("UserInfoResult", UserInfoSerializer()),
    )
    @action(methods=["get"], detail=False, url_path="me")
    def me(self, request, *args, **kwargs):
        """获取当前登录用户基础信息。"""
        serializer = UserInfoSerializer(request.user)
        return success(serializer.data)

    @permission_required(["sys:user:update"])
    @extend_schema(
        summary="获取用户信息",
        responses=resp("UserFormResult", UserFormSerializer()),
    )
    @action(methods=["get"], detail=True, url_path="form")
    def form(self, request, pk=None, *args, **kwargs):
        """获取单个用户表单信息（编辑页回显）。"""
        user = self.get_object()
        serializer = UserFormSerializer(user)
        return success(serializer.data, msg="获取用户信息成功")

    @permission_required(["sys:user:reset-password"])
    @extend_schema(
        summary="重置用户密码",
        parameters=[
            OpenApiParameter(
                name="password",
                location=OpenApiParameter.QUERY,
                required=True,
                type=str,
                description="新密码（通过 query 参数传入）",
            ),
        ],
        responses=resp("UserPasswordResetResult", serializers.DictField(required=False)),
    )
    @action(methods=["put"], detail=True, url_path=r"password/reset")
    def password_reset(self, request, pk=None, *args, **kwargs):
        """重置指定用户密码。"""
        user = self.get_object()

        if not _check_user_data_permission(request=request, operation="update", target_user_id=user.id):
            raise PermissionDenied("您没有权限重置该用户密码")

        password = request.query_params.get('password')
        if not password:
            return error("密码不能为空")

        if user.username == 'root' and request.user.username != 'root':
            raise PermissionDenied("root 用户不允许重置密码")

        serializer = UserPasswordResetSerializer(user, data={'password': password})
        if serializer.is_valid():
            serializer.save()
            invalidate_user_sessions(user.id)
            return success({}, msg="密码重置成功")
        return error(serializer.errors)

    @permission_required(["sys:user:update"])
    @extend_schema(
        summary="更新用户状态",
        parameters=[
            OpenApiParameter(
                name="status",
                location=OpenApiParameter.QUERY,
                required=True,
                type=int,
                description="用户状态：1 启用 / 0 禁用（通过 query 参数传入）",
            ),
        ],
        responses=resp("UserStatusUpdateResult", serializers.DictField(required=False)),
    )
    @action(methods=["patch"], detail=True, url_path="status")
    def status(self, request, pk=None, *args, **kwargs):
        """启用/禁用用户。"""
        user = self.get_object()
        try:
            status_value = int(request.query_params.get('status', ''))
        except (ValueError, TypeError):
            return error("状态值格式不正确")

        if not _check_user_data_permission(request=request, operation="update", target_user_id=user.id):
            raise PermissionDenied("您没有权限更新该用户状态")

        if user.id == request.user.id and status_value == 0:
            return error("您不能禁用自己")
        if user.username == 'root' and status_value == 0:
            return error("您不能禁用 root 用户")

        serializer = UserStatusSerializer(user, data={'status': status_value})
        if serializer.is_valid():
            serializer.save()
            if status_value == 0:
                invalidate_user_sessions(user.id)
            status_text = "启用" if status_value == 1 else "禁用"
            return success({}, msg=f"用户 {status_text} 成功")
        return error(serializer.errors)

    @log_api('获取用户个人信息', module='用户模块')
    @extend_schema(
        summary="获取用户个人信息",
        responses=resp(
            "UserProfileResult",
            inline_serializer(
                name="UserProfileData",
                fields={
                    "id": serializers.CharField(),
                    "username": serializers.CharField(),
                    "nickname": serializers.CharField(allow_blank=True, required=False),
                    "avatar": serializers.CharField(allow_blank=True, required=False),
                    "gender": serializers.IntegerField(),
                    "mobile": serializers.CharField(allow_blank=True, required=False),
                    "email": serializers.CharField(allow_blank=True, required=False),
                    "deptName": serializers.CharField(allow_blank=True, required=False),
                    "roleNames": serializers.CharField(allow_blank=True, required=False),
                    "createTime": serializers.DateTimeField(),
                },
            ),
        ),
    )
    @action(methods=["get"], detail=False, url_path="profile")
    def profile(self, request, *args, **kwargs):
        """获取当前用户个人资料（包含部门名、角色名等聚合字段）。"""
        user = request.user
        dept_name = user.dept.name if user.dept else ""
        roles = user.roles.all()
        role_names = ", ".join([role.name for role in roles]) if roles else ""
        data = {
            "id": str(user.id),
            "username": user.username,
            "nickname": user.nickname,
            "avatar": user.avatar,
            "gender": user.gender,
            "mobile": user.mobile,
            "email": user.email,
            "deptName": dept_name,
            "roleNames": role_names,
            "createTime": user.date_joined,
        }
        return success(data)

    @log_api('更新用户个人信息', module='用户模块')
    @extend_schema(
        summary="更新用户个人信息",
        request=UserProfileUpdateSerializer,
        responses=resp("UserProfileUpdateResult", serializers.BooleanField()),
    )
    @profile.mapping.put
    def profile_update(self, request, *args, **kwargs):
        """更新当前用户个人资料（partial update）。"""
        user = request.user
        serializer = UserProfileUpdateSerializer(user, data=request.data, partial=True)
        if not serializer.is_valid():
            return error(serializer.errors)

        if not serializer.validated_data:
            return error("请至少修改一项")

        serializer.save()
        return success(True, msg="更新用户个人信息成功")

    @log_api('修改用户密码', module='用户模块')
    @extend_schema(
        summary="修改用户密码",
        request=UserPasswordChangeSerializer,
        responses=resp("UserPasswordChangeResult", serializers.BooleanField()),
    )
    @action(methods=["put"], detail=False, url_path="password")
    def password_change(self, request, *args, **kwargs):
        """修改当前用户密码。"""
        serializer = UserPasswordChangeSerializer(
            request.user,
            data=request.data,
            context={'request': request},
        )
        if serializer.is_valid():
            serializer.save()
            invalidate_user_sessions(request.user.id)
            return success(True, msg="修改密码成功")
        return error(serializer.errors)

    @extend_schema(
        summary="发送邮箱验证码",
        parameters=[
            OpenApiParameter(
                name="email",
                location=OpenApiParameter.QUERY,
                required=True,
                type=str,
                description="邮箱地址",
            ),
        ],
        responses=resp("UserEmailCodeResult", serializers.BooleanField()),
    )
    @action(methods=["post"], detail=False, url_path=r"email/code")
    def email_code(self, request, *args, **kwargs):
        """发送邮箱验证码（用于后续绑定/修改邮箱）。"""
        email = request.query_params.get('email')
        if not email:
            return error('邮箱不能为空')
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            return error('邮箱格式不正确')

        result = generate_and_send_email_code(
            email=email,
            subject='邮箱验证码',
            user_id=request.user.id,
        )
        if result.get('success'):
            return success(True, msg='验证码发送成功')
        return error('验证码发送失败', status=500)

    @extend_schema(
        summary="更新用户邮箱",
        request=UserEmailUpdateSerializer,
        responses=resp("UserEmailUpdateResult", serializers.BooleanField()),
    )
    @action(methods=["put"], detail=False, url_path="email")
    def email_update(self, request, *args, **kwargs):
        """绑定或更换邮箱（需要前置验证码校验，具体逻辑在 serializer 内）。"""
        serializer = UserEmailUpdateSerializer(
            request.user,
            data=request.data,
            context={'request': request},
        )
        if serializer.is_valid():
            serializer.save()
            return success(True, msg='更新邮箱成功')
        return error(serializer.errors)

    @extend_schema(
        summary="解绑邮箱",
        request=PasswordVerifySerializer,
        responses=resp("UserEmailUnbindResult", serializers.BooleanField()),
    )
    @email_update.mapping.delete
    def email_unbind(self, request, *args, **kwargs):
        """解绑邮箱（需要校验当前密码）。"""
        serializer = PasswordVerifySerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return error(serializer.errors)

        user = request.user
        if not user.email:
            return error("当前账号未绑定邮箱")

        if not user.check_password(serializer.validated_data.get("password")):
            return error({"password": "当前密码错误"})

        user.email = None
        user.save(update_fields=["email"])
        return success(True, msg="解绑邮箱成功")

    @ip_rate_limit(60)
    @extend_schema(
        summary="发送手机验证码",
        parameters=[
            OpenApiParameter(
                name="mobile",
                location=OpenApiParameter.QUERY,
                required=True,
                type=str,
                description="手机号",
            ),
        ],
        responses=resp("UserMobileCodeResult", serializers.BooleanField()),
    )
    @action(methods=["post"], detail=False, url_path=r"mobile/code")
    def mobile_code(self, request, *args, **kwargs):
        """发送手机验证码（限流由 `@ip_rate_limit` 控制）。"""
        mobile = request.query_params.get('mobile')
        if not mobile:
            return error('手机号不能为空')
        if not re.match(r'^1[3-9]\d{9}$', mobile):
            return error('手机号格式不正确')

        result = send_mobile_code(mobile, request.user.id)
        if result.get('success'):
            return success(True, msg='验证码发送成功')
        return error(result.get('message') or '验证码发送失败', status=500)

    @extend_schema(
        summary="绑定或更换手机号",
        request=UserMobileUpdateSerializer,
        responses=resp("UserMobileUpdateResult", serializers.BooleanField()),
    )
    @action(methods=["put"], detail=False, url_path="mobile")
    def mobile_update(self, request, *args, **kwargs):
        """绑定或更换手机号（需要前置验证码校验，具体逻辑在 serializer 内）。"""
        serializer = UserMobileUpdateSerializer(
            request.user,
            data=request.data,
            context={'request': request},
        )
        if serializer.is_valid():
            serializer.save()
            return success(True, msg='更新手机号成功')
        return error(serializer.errors)

    @extend_schema(
        summary="解绑手机号",
        request=PasswordVerifySerializer,
        responses=resp("UserMobileUnbindResult", serializers.BooleanField()),
    )
    @mobile_update.mapping.delete
    def mobile_unbind(self, request, *args, **kwargs):
        """解绑手机号（需要校验当前密码）。"""
        serializer = PasswordVerifySerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return error(serializer.errors)

        user = request.user
        if not user.mobile:
            return error("当前账号未绑定手机号")

        if not user.check_password(serializer.validated_data.get("password")):
            return error({"password": "当前密码错误"})

        user.mobile = None
        user.save(update_fields=["mobile"])
        return success(True, msg="解绑手机号成功")

    @permission_required(["sys:user:import"])
    @extend_schema(
        summary="下载用户导入模板",
        responses={200: OpenApiTypes.BINARY},
    )
    @action(methods=["get"], detail=False, url_path="template")
    def template(self, request, *args, **kwargs):
        """下载用户导入模板（Excel）。"""
        from urllib.parse import quote

        excel_bytes = build_user_import_template_excel_bytes()
        response = HttpResponse(
            excel_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = quote("用户导入.xlsx")
        response['Content-Disposition'] = f"attachment; filename={filename}"
        return response

    @permission_required(["sys:user:import"])
    @extend_schema(
        summary="导入用户",
        request=inline_serializer(
            name="UserImportRequest",
            fields={
                "file": serializers.FileField(),
            },
        ),
        responses=resp(
            "UserImportResult",
            inline_serializer(
                name="UserImportData",
                fields={
                    "validCount": serializers.IntegerField(),
                    "invalidCount": serializers.IntegerField(),
                    "messageList": serializers.ListField(child=serializers.CharField()),
                },
            ),
        ),
    )
    @action(methods=["post"], detail=False, url_path="import")
    def import_users(self, request, *args, **kwargs):
        """批量导入用户（Excel）。

        - 仅支持 `.xlsx/.xls`。
        - 业务校验与逐行错误信息由 `import_users_from_excel` 生成。
        """
        if 'file' not in request.FILES:
            return error('请上传文件')
        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return error('只支持 .xlsx/.xls 文件')

        result = import_users_from_excel(file=file)
        return success(result["data"], msg=result["msg"])

    @permission_required(["sys:user:export"])
    @extend_schema(
        summary="导出用户",
        parameters=[
            OpenApiParameter(
                name="keywords",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="关键词（用户名/昵称/手机号）",
            ),
            OpenApiParameter(
                name="status",
                location=OpenApiParameter.QUERY,
                required=False,
                type=int,
                description="用户状态：1 启用 / 0 禁用",
            ),
            OpenApiParameter(
                name="deptId",
                location=OpenApiParameter.QUERY,
                required=False,
                type=int,
                description="部门ID",
            ),
            OpenApiParameter(
                name="roleIds",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="角色ID列表，多个以英文逗号(,)分割",
            ),
            OpenApiParameter(
                name="createTime",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="创建时间范围：start,end（可被 parse_datetime 解析）",
            ),
            OpenApiParameter(
                name="field",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="排序字段（默认 id）",
            ),
            OpenApiParameter(
                name="direction",
                location=OpenApiParameter.QUERY,
                required=False,
                type=str,
                description="排序方向：ASC / DESC（默认 DESC）",
            ),
        ],
        responses={200: OpenApiTypes.BINARY},
    )
    @action(methods=["get"], detail=False, url_path="export")
    def export(self, request, *args, **kwargs):
        """导出用户列表（Excel 附件）。"""
        keywords = request.query_params.get('keywords', '')
        status = request.query_params.get('status')
        dept_id = request.query_params.get('deptId')
        role_ids = request.query_params.get('roleIds', '')
        create_time = request.query_params.get('createTime', '')
        field = request.query_params.get('field', 'id')
        direction = request.query_params.get('direction', 'DESC')

        excel_bytes = export_users_to_excel_bytes(
            keywords=keywords,
            status=status,
            dept_id=dept_id,
            role_ids=role_ids,
            create_time=create_time,
            field=field,
            direction=direction,
        )

        from urllib.parse import quote

        response = HttpResponse(
            excel_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = quote("用户列表.xlsx")
        response['Content-Disposition'] = f"attachment; filename={filename}"
        return response

    @permission_required(["sys:user:list"])
    @extend_schema(
        summary="用户下拉选项",
        responses=resp("UserOptionsResult", serializers.ListField(child=serializers.DictField())),
    )
    @action(methods=["get"], detail=False, url_path="options")
    def options(self, request, *args, **kwargs):
        """用户下拉选项（仅返回启用且未删除用户；受数据权限约束）。"""
        data_permission = data_permission_required(request.user, "select")

        query = Q(is_deleted=0, status=1)
        if isinstance(data_permission, list) and data_permission:
            query &= Q(id__in=data_permission)

        users = User.objects.filter(query).order_by("id")
        result = [{"value": str(u.id), "label": (u.nickname or u.username)} for u in users]
        return success(result)
